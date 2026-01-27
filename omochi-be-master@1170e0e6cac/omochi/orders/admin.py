from django import forms
from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.shortcuts import redirect
from django.utils.html import format_html
from openpyxl.styles import Alignment
from omochi.coupons.models import UserCoupon
from .models import Order, OrderItem, OrderStatusHistory, OrderItemsMerged, OrderQuestion
from io import BytesIO
import openpyxl
import pytz
from django.utils import timezone


@admin.action(description='Export to Excel')
def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Items"

    headers = [
        'Date of Order',
        'Order ID',
        'Reservation ID',
        'Order Code',
        'Reservation Code',
        'User',
        'User ID',
        'Order Type',
        'Payment Method',
        'Venue Name',
        'Venue ID',
        'Order Link',
        'Ref Source User ID',
        'Time Slot',
        'Party Size',
        'Status',
        'Menu Item',
        'Menu Unit Price',
        'Quantity',
        'Subtotal Price',
        'Application Fee Amount',
        'Takeout Fee Subsidized Amount'
    ]
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # JST timezone
    jst = pytz.timezone('Asia/Tokyo')

    for obj in queryset:
        # Get the referral information
        from omochi.ref_logs.models import RefLog
        from django.conf import settings
        from django.apps import apps

        ref_link = 'None'
        ref_source_user_id = None
        
        try:
            # Find the ref_log where action_id matches the order id and action_type is 'ORDER'
            ref_log = RefLog.objects.filter(action_id=str(obj.order.id), action_type='ORDER').first()
            if ref_log and ref_log.ref_id:
                ref_link = ref_log.ref_id
                # Get the user who owns this referral code
                User = apps.get_model(settings.AUTH_USER_MODEL)
                referrer = User.objects.filter(ref_code=ref_log.ref_id).first()
                if referrer:
                    ref_source_user_id = referrer.id
        except Exception as e:
            ref_link = f'Error: {str(e)}'

        # Get menu price
        menu_price = 'N/A'
        try:
            if hasattr(obj.menu_item, 'price'):
                # Return take_out_price if order type is TAKEOUT and take_out_price exists
                if obj.order.order_type == 'TAKEOUT' and hasattr(obj.menu_item, 'take_out_price') and obj.menu_item.take_out_price is not None:
                    menu_price = obj.menu_item.take_out_price
                else:
                    menu_price = obj.menu_item.price
        except:
            pass

        # Get time slot
        time_slot = 'N/A'
        try:
            if hasattr(obj.order, 'time_slot') and obj.order.time_slot:
                time_slot = f"{obj.order.time_slot.start_time} - {obj.order.time_slot.end_time}"
            elif obj.order.start_time and obj.order.end_time:
                time_slot = f"{obj.order.start_time} - {obj.order.end_time}"
            else:
                time_slot = 'N/A'
        except:
            pass

        # Convert order date to JST
        order_date_jst = 'N/A'
        if hasattr(obj.order, 'order_date') and obj.order.order_date:
            # Convert UTC to JST
            utc_time = timezone.localtime(obj.order.order_date, pytz.UTC)
            jst_time = utc_time.astimezone(jst)
            order_date_jst = jst_time.strftime('%Y-%m-%d %H:%M:%S')

        # Get party size with conditional logic order_type
        party_size = 'N/A'
        if hasattr(obj.order, 'order_type'):
            if obj.order.order_type == 'DINE_IN':
                party_size = obj.order.party_size if hasattr(obj.order, 'party_size') else 'N/A'
            elif obj.order.order_type == 'TAKEOUT':
                party_size = 1

        # Get reservation data
        reservation_code = 'N/A'
        reservation_id = 'N/A'
        if hasattr(obj.order, 'reservation') and obj.order.reservation:
            reservation_code = obj.order.reservation.reservation_code or 'N/A'
            reservation_id = str(obj.order.reservation.id) or 'N/A'

        # Create the row and convert all values to string to avoid Excel conversion issues
        row = [
            order_date_jst,
            str(obj.order.id),
            str(reservation_id),
            str(obj.order.order_code) if hasattr(obj.order, 'order_code') and obj.order.order_code else 'N/A',
            str(reservation_code),
            str(obj.order.user) if obj.order.user else 'N/A',
            str(obj.order.user.id) if obj.order.user else 'N/A',
            str(obj.order.order_type) if hasattr(obj.order, 'order_type') else 'N/A',
            str(obj.order.payment_method) if hasattr(obj.order, 'payment_method') else 'N/A',
            str(obj.order.venue),
            str(obj.order.venue.id) if obj.order.venue else 'N/A',
            str(ref_link),
            str(ref_source_user_id) if ref_source_user_id is not None else 'N/A',
            str(time_slot),
            str(party_size),
            str(obj.order.status) if hasattr(obj.order, 'status') and obj.order.status else 'N/A',
            str(obj.menu_item.name),
            str(menu_price),
            str(obj.quantity),
            str(obj.subtotal),
            str(obj.order.application_fee_amount) if hasattr(obj.order, 'application_fee_amount') and obj.order.application_fee_amount is not None else 'N/A',
            str(obj.order.takeout_fee_subsidized_amount) if hasattr(obj.order, 'takeout_fee_subsidized_amount') and obj.order.takeout_fee_subsidized_amount is not None else 'N/A',
        ]
        
        # Ensure all values in the row are Excel-compatible
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        content=buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=order_items.xlsx'

    return response


@admin.action(description='Export Order Items Merged to Excel')
def export_unified_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Items Merged"

    headers = [
        'Record Type',
        'Date of Order',
        'Order ID',
        'Reservation ID',
        'Order Code',
        'Reservation Code',
        'User',
        'User ID',
        'Order Type',
        'Payment Method',
        'Venue Name',
        'Venue ID',
        'Order Link',
        'Ref Source User ID',
        'Status',
        'Time Slot',
        'Party Size',
        'Menu Item',
        'Menu Unit Price',
        'Quantity',
        'Subtotal',
        'Total Amount',
        'Application Fee Amount',
        'Takeout Fee Subsidized Amount'
    ]
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # JST timezone
    jst = pytz.timezone('Asia/Tokyo')

    for obj in queryset:
        # Convert date to JST - order_date now handles both order_date and reservation created_at
        order_date_jst = 'N/A'
        if obj.order_date:
            # Convert UTC to JST
            utc_time = timezone.localtime(obj.order_date, pytz.UTC)
            jst_time = utc_time.astimezone(jst)
            order_date_jst = jst_time.strftime('%Y-%m-%d %H:%M:%S')

        # Format time slot
        time_slot = 'N/A'
        if obj.start_time and obj.end_time:
            time_slot = f"{obj.start_time} - {obj.end_time}"

        # Get party size - use unified logic since reservations now have order_type = 'DINE_IN'
        party_size = 'N/A'
        try:
            if obj.order_type == 'DINE_IN':
                party_size = obj.party_size if obj.party_size is not None else 'N/A'
            elif obj.order_type == 'TAKEOUT':
                party_size = 1
            else:
                # Fallback for any edge cases
                party_size = obj.party_size if obj.party_size is not None else 'N/A'
        except Exception:
            party_size = 'N/A'

        # Get venue name - use direct field from view first, fallback to query
        venue_name = 'N/A'
        try:
            if obj.venue_name:
                venue_name = obj.venue_name
            elif obj.venue_id:
                from omochi.venues.models import Venue
                venue = Venue.objects.filter(id=obj.venue_id).first()
                venue_name = venue.name if venue else 'N/A'
        except Exception:
            pass

        # Get user display - use direct user_email field from view first, fallback to query
        user_display = 'N/A'
        try:
            if obj.user_email:
                user_display = obj.user_email
            elif obj.user_id:
                from django.conf import settings
                from django.apps import apps
                User = apps.get_model(settings.AUTH_USER_MODEL)
                user = User.objects.filter(id=obj.user_id).first()
                user_display = str(user) if user else 'N/A'
        except Exception:
            pass

        # Get order link
        order_link = 'None'
        try:
            from omochi.ref_logs.models import RefLog
            action_id = obj.order_id if obj.order_id else obj.reservation_id
            action_type = 'ORDER' if obj.order_id else 'RESERVATION'
            
            if action_id:
                ref_log = RefLog.objects.filter(action_id=str(action_id), action_type=action_type).first()
                if ref_log:
                    order_link = ref_log.ref_id
        except Exception as e:
            order_link = f'Error: {str(e)}'

        # Get referral source user ID
        ref_source_user_id = 'None'
        try:
            from omochi.ref_logs.models import RefLog
            from django.conf import settings
            from django.apps import apps
            
            action_id = obj.order_id if obj.order_id else obj.reservation_id
            action_type = 'ORDER' if obj.order_id else 'RESERVATION'
            
            if action_id:
                ref_log = RefLog.objects.filter(action_id=str(action_id), action_type=action_type).first()
                if ref_log and ref_log.ref_id:
                    User = apps.get_model(settings.AUTH_USER_MODEL)
                    referrer = User.objects.filter(ref_code=ref_log.ref_id).first()
                    if referrer:
                        ref_source_user_id = str(referrer.id)
        except Exception as e:
            ref_source_user_id = f'Error: {str(e)}'

        # Get menu item name
        menu_item_name = 'N/A'
        try:
            if obj.menu_item_id:
                from omochi.menus.models import MenuItem
                menu_item = MenuItem.objects.filter(id=obj.menu_item_id).first()
                menu_item_name = menu_item.name if menu_item else 'N/A'
        except Exception:
            pass

        # Get menu price
        menu_price = 'N/A'
        try:
            if obj.menu_item_id:
                from omochi.menus.models import MenuItem
                menu_item = MenuItem.objects.filter(id=obj.menu_item_id).first()
                if menu_item:
                    if obj.order_type == 'TAKEOUT' and hasattr(menu_item, 'take_out_price') and menu_item.take_out_price is not None:
                        menu_price = menu_item.take_out_price
                    else:
                        menu_price = menu_item.price
        except Exception:
            pass

        # Create the row
        row = [
            str(obj.record_type),
            str(order_date_jst),
            str(obj.order_id) if obj.order_id else 'N/A',
            str(obj.reservation_id) if obj.reservation_id else 'N/A',
            str(obj.order_code) if obj.order_code else 'N/A',
            str(obj.reservation_code) if obj.reservation_code else 'N/A',
            str(user_display),
            str(obj.user_id) if obj.user_id else 'N/A',
            str(obj.order_type) if obj.order_type else 'N/A',
            str(obj.payment_method) if obj.payment_method else 'N/A',
            str(venue_name),
            str(obj.venue_id) if obj.venue_id else 'N/A',
            str(order_link),
            str(ref_source_user_id),
            str(obj.status) if obj.status else 'N/A',
            str(time_slot),
            str(party_size),
            str(menu_item_name),
            str(menu_price),
            str(obj.quantity) if obj.quantity else 'N/A',
            str(obj.subtotal) if obj.subtotal else 'N/A',
            str(obj.total_amount) if obj.total_amount else 'N/A',
            str(obj.application_fee_amount) if obj.application_fee_amount else 'N/A',
            str(obj.takeout_fee_subsidized_amount) if obj.takeout_fee_subsidized_amount else 'N/A',
        ]
        
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        content=buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=order_items_merged.xlsx'

    return response


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ('id',)


class OrderStatusHistoryInline(admin.TabularInline):
    model = OrderStatusHistory
    extra = 0
    readonly_fields = (
        'id',
        'old_status',
        'new_status',
        'changed_at',
        'changed_by',
    )
    can_delete = False


class UserCouponInline(admin.TabularInline):
    model = UserCoupon
    extra = 0
    readonly_fields = ('id', 'created_at', 'updated_at')
    can_delete = False


class OrderQuestionInline(admin.TabularInline):
    model = OrderQuestion
    extra = 0
    readonly_fields = ('id', 'created_at', 'updated_at')
    fields = ('order_index', 'question', 'question_en', 'answer', 'answer_en',)
    ordering = ('order_index',)
    formfield_overrides = {
        model._meta.get_field('question').__class__: {
            'widget': forms.TextInput(attrs={'size': 25, 'style': 'min-width:150px;'})
        },
        model._meta.get_field('question_en').__class__: {
            'widget': forms.TextInput(attrs={'size': 25, 'style': 'min-width:150px;'})
        },
        model._meta.get_field('answer').__class__: {
            'widget': admin.widgets.AdminTextareaWidget(
                attrs={
                    'rows': 3,
                    'cols': 15,
                    'style': 'resize:vertical; min-width:200px; width:100%;'
                }
            )
        },
        model._meta.get_field('answer_en').__class__: {
            'widget': admin.widgets.AdminTextareaWidget(
                attrs={
                    'rows': 3,
                    'cols': 15,
                    'style': 'resize:vertical; min-width:200px; width:100%;'
                }
            )
        },
    }


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'get_order_date_jst',
        'user',
        'get_user_id',
        'order_type',
        'payment_method',
        'venue',
        'venue_id',
        'get_order_link_type',
        'get_ref_source_user',
        'get_time_slot',
        'status',
        'payment_status',
        'get_total_item_amount',
        'application_fee_amount',
        'takeout_fee_subsidized_amount',
        'order_discount_amount',
        'application_fee_discount_amount',
        'total',
    )
    list_filter = ('status', 'order_type', 'payment_status', 'venue')
    search_fields = ('id', 'user__username', 'user__email', 'venue__name')
    readonly_fields = ('id', 'order_date', 'updated_at')
    inlines = [OrderItemInline, OrderStatusHistoryInline, UserCouponInline, OrderQuestionInline]
    date_hierarchy = 'order_date'
    ordering = ('-order_date',)  # Order by order date descending (most recent first)
    
    def get_queryset(self, request):
        """Optimize queryset with prefetch_related for better performance"""
        return super().get_queryset(request).prefetch_related('order_questions')
    
    def get_order_date_jst(self, obj):
        # JST timezone
        jst = pytz.timezone('Asia/Tokyo')
        
        if obj.order_date:
            # Convert UTC to JST
            utc_time = timezone.localtime(obj.order_date, pytz.UTC)
            jst_time = utc_time.astimezone(jst)
            return jst_time.strftime('%Y-%m-%d %H:%M:%S')
        return 'N/A'
    
    get_order_date_jst.short_description = 'Order Date (JST)'
    
    def get_user_id(self, obj):
        return obj.user.id if obj.user else None
    
    get_user_id.short_description = 'User ID'
    
    def get_order_link_type(self, obj):
        from omochi.ref_logs.models import RefLog
        try:
            # Find the ref_log where action_id matches the order id and action_type is 'ORDER'
            ref_log = RefLog.objects.filter(action_id=str(obj.id), action_type='ORDER').first()
            if ref_log:
                # If ref_log exists, return the ref_id (referral code)
                return ref_log.ref_id
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_order_link_type.short_description = 'Order Link'
    
    def get_ref_source_user(self, obj):
        from omochi.ref_logs.models import RefLog
        from django.conf import settings
        from django.apps import apps
        
        try:
            # Find the ref_log where action_id matches the order id and action_type is 'ORDER'
            ref_log = RefLog.objects.filter(action_id=str(obj.id), action_type='ORDER').first()
            if ref_log and ref_log.ref_id:
                # Get the user who owns this referral code
                User = apps.get_model(settings.AUTH_USER_MODEL)
                referrer = User.objects.filter(ref_code=ref_log.ref_id).first()
                if referrer:
                    return str(referrer.id)
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_ref_source_user.short_description = 'Ref Source User ID'
    
    def get_time_slot(self, obj):
        try:
            if hasattr(obj, 'time_slot') and obj.time_slot:
                return f"{obj.time_slot.start_time} - {obj.time_slot.end_time}"
            elif obj.start_time and obj.end_time:
                return f"{obj.start_time} - {obj.end_time}"
            else:
                return 'N/A'
        except:
            return 'N/A'
    
    get_time_slot.short_description = 'Time Slot'
    
    def get_total_item_amount(self, obj):
        return obj.total_amount
    
    get_total_item_amount.short_description = 'Total Item Amount'


@admin.register(OrderItem)
class OrderItemAdmin(admin.ModelAdmin):
    list_display = (
        'get_order_date',
        'order__id',
        'get_reservation_id',
        'get_order_code',
        'get_reservation_code',
        'get_user',
        'get_user_id',
        'get_order_type',
        'get_payment_method',
        'get_venue',
        'get_venue_id',
        'get_order_link',
        'get_ref_source_user_id',
        'get_time_slot',
        'get_party_size',
        'get_order_status',
        'menu_item__name',
        'get_menu_price',
        'quantity',
        'subtotal',
        'get_application_fee_amount',
        'get_takeout_fee_subsidized_amount',
    )
    list_filter = ('order__status', 'order__order_type', 'order__payment_status', 'order__venue')
    search_fields = ('order__id', 'menu_item__name', 'order__venue__name')
    readonly_fields = ('id',)
    actions = [export_to_excel]
    ordering = ('-order__order_date',)  # Order by order date descending (most recent first)

    def get_order_date(self, obj):
        # JST timezone
        jst = pytz.timezone('Asia/Tokyo')
        
        if hasattr(obj.order, 'order_date') and obj.order.order_date:
            # Convert UTC to JST
            utc_time = timezone.localtime(obj.order.order_date, pytz.UTC)
            jst_time = utc_time.astimezone(jst)
            return jst_time.strftime('%Y-%m-%d %H:%M:%S')
        return 'N/A'
    
    get_order_date.short_description = 'Date of Order'
    
    def get_reservation_id(self, obj):
        if hasattr(obj.order, 'reservation') and obj.order.reservation:
            return obj.order.reservation.id or 'N/A'
        return 'N/A'
    
    get_reservation_id.short_description = 'Reservation ID'

    def get_order_code(self, obj):
        return obj.order.order_code if hasattr(obj.order, 'order_code') and obj.order.order_code else 'N/A'
    
    get_order_code.short_description = 'Order Code'

    def get_reservation_code(self, obj):
        if hasattr(obj.order, 'reservation') and obj.order.reservation:
            return obj.order.reservation.reservation_code or 'N/A'
        return 'N/A'
    
    get_reservation_code.short_description = 'Reservation Code'

    def get_user_id(self, obj):
        return obj.order.user.id if obj.order.user else 'N/A'
    
    get_user_id.short_description = 'User ID'
    
    def get_user(self, obj):
        return obj.order.user
    
    get_user.short_description = 'User'
    
    def get_order_type(self, obj):
        return str(obj.order.order_type) if hasattr(obj.order, 'order_type') else 'N/A'
    
    get_order_type.short_description = 'Order Type'
    
    def get_payment_method(self, obj):
        return str(obj.order.payment_method) if hasattr(obj.order, 'payment_method') else 'N/A'
    
    get_payment_method.short_description = 'Payment Method'

    def get_venue(self, obj):
        return obj.order.venue

    get_venue.short_description = 'Venue Name'
    
    def get_venue_id(self, obj):
        return obj.order.venue.id if obj.order.venue else 'N/A'
    
    get_venue_id.short_description = 'Venue ID'
    
    def get_order_link(self, obj):
        from omochi.ref_logs.models import RefLog
        try:
            # Find the ref_log where action_id matches the order id and action_type is 'ORDER'
            ref_log = RefLog.objects.filter(action_id=str(obj.order.id), action_type='ORDER').first()
            if ref_log:
                # If ref_log exists, return the ref_id (referral code)
                return ref_log.ref_id
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_order_link.short_description = 'Order Link'
    
    def get_ref_source_user_id(self, obj):
        from omochi.ref_logs.models import RefLog
        from django.conf import settings
        from django.apps import apps
        
        try:
            # Find the ref_log where action_id matches the order id and action_type is 'ORDER'
            ref_log = RefLog.objects.filter(action_id=str(obj.order.id), action_type='ORDER').first()
            if ref_log and ref_log.ref_id:
                # Get the user who owns this referral code
                User = apps.get_model(settings.AUTH_USER_MODEL)
                referrer = User.objects.filter(ref_code=ref_log.ref_id).first()
                if referrer:
                    return str(referrer.id)
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_ref_source_user_id.short_description = 'Ref Source User ID'

    def get_time_slot(self, obj):
        try:
            if hasattr(obj.order, 'time_slot') and obj.order.time_slot:
                return f"{obj.order.time_slot.start_time} - {obj.order.time_slot.end_time}"
            elif obj.order.start_time and obj.order.end_time:
                return f"{obj.order.start_time} - {obj.order.end_time}"
            return 'N/A'
        except:
            return 'N/A'

    get_time_slot.short_description = 'Time Slot'
    
    def get_party_size(self, obj):
        if hasattr(obj.order, 'order_type'):
            if obj.order.order_type == 'DINE_IN':
                return obj.order.party_size if hasattr(obj.order, 'party_size') else 'N/A'
            elif obj.order.order_type == 'TAKEOUT':
                return 1
        return 'N/A'
    
    get_party_size.short_description = 'Party Size'

    def get_order_status(self, obj):
        return obj.order.status if hasattr(obj.order, 'status') and obj.order.status else 'N/A'
    
    get_order_status.short_description = 'Status'
    
    def get_menu_price(self, obj):
        try:
            if hasattr(obj.menu_item, 'price'):
                # Return take_out_price if order type is TAKEOUT and take_out_price exists
                if obj.order.order_type == 'TAKEOUT' and hasattr(obj.menu_item, 'take_out_price') and obj.menu_item.take_out_price is not None:
                    return obj.menu_item.take_out_price
                else:
                    return obj.menu_item.price
            else:
                return 'N/A'
        except:
            return 'N/A'
            
    get_menu_price.short_description = 'Unit Price'

    def get_application_fee_amount(self, obj):
        return obj.order.application_fee_amount if hasattr(obj.order, 'application_fee_amount') and obj.order.application_fee_amount is not None else 'N/A'
    
    get_application_fee_amount.short_description = 'Application Fee Amount'
    
    def get_takeout_fee_subsidized_amount(self, obj):
        return obj.order.takeout_fee_subsidized_amount if hasattr(obj.order, 'takeout_fee_subsidized_amount') and obj.order.takeout_fee_subsidized_amount is not None else 'N/A'
    
    get_takeout_fee_subsidized_amount.short_description = 'Takeout Fee Subsidized Amount'
    
@admin.register(OrderStatusHistory)
class OrderStatusHistoryAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'order',
        'old_status',
        'new_status',
        'changed_at',
        'changed_by',
    )
    list_filter = ('old_status', 'new_status')
    search_fields = ('order__id',)
    readonly_fields = (
        'id',
        'order',
        'old_status',
        'new_status',
        'changed_at',
        'changed_by',
    )
    date_hierarchy = 'changed_at'


@admin.register(OrderItemsMerged)
class OrderItemsMergedAdmin(admin.ModelAdmin):
    list_display = (
        'get_record_type',
        'get_order_date_jst',
        'order_id',
        'reservation_id',
        'order_code',
        'reservation_code',
        'get_user',
        'user_id',
        'order_type',
        'payment_method',
        'get_venue_name',
        'venue_id',
        'get_order_link',
        'get_ref_source_user_id',
        'status',
        'get_time_slot',
        'get_party_size',
        'get_menu_item_name',
        'get_menu_price',
        'quantity',
        'subtotal',
        'total_amount',
        'application_fee_amount',
        'takeout_fee_subsidized_amount',
    )
    
    def get_list_display_links(self, request, list_display):
        return ('get_record_type',)
    
    list_filter = ('record_type', 'order_type', 'status', 'payment_method')
    search_fields = ('order_id', 'reservation_id', 'order_code', 'reservation_code', 'user_id')
    readonly_fields = (
        'id', 'order_item_id', 'order_id', 'reservation_id', 'order_code', 'reservation_code',
        'order_date', 'user_id', 'venue_id', 'order_type', 'payment_method', 'status',
        'party_size', 'time_slot_id', 'start_time', 'end_time', 'total_amount',
        'application_fee_amount', 'takeout_fee_subsidized_amount', 'menu_item_id',
        'quantity', 'subtotal', 'special_request', 'record_type'
    )
    actions = [export_unified_to_excel]
    ordering = ('-order_date',)

    def has_add_permission(self, request):
        return False  # Cannot add to a view

    def has_change_permission(self, request, obj=None):
        return False  # Cannot change a view

    def has_delete_permission(self, request, obj=None):
        return False  # Cannot delete from a view
    
    def changelist_view(self, request, extra_context=None):
        """Override to add custom template with row linking JavaScript"""
        extra_context = extra_context or {}
        return super().changelist_view(request, extra_context)
    
    def response_change(self, request, obj):
        """Override to redirect to appropriate admin pages based on record type"""
        if obj.record_type == 'Order' and obj.order_item_id:
            # Redirect to order item change page
            return redirect(reverse('admin:orders_orderitem_change', args=[obj.order_item_id]))
        elif obj.record_type == 'Reservation' and obj.reservation_id:
            # Redirect to reservation change page
            return redirect(reverse('admin:reservations_reservation_change', args=[obj.reservation_id]))
        return super().response_change(request, obj)
    
    def get_urls(self):
        """Override URLs to redirect change pages to appropriate models"""
        from django.urls import path
        urls = super().get_urls()
        
        # Add custom URL pattern that redirects to appropriate change page
        custom_urls = [
            path('<path:object_id>/change/', self.redirect_to_appropriate_change, name='%s_%s_change' % (self.model._meta.app_label, self.model._meta.model_name)),
        ]
        return custom_urls + urls
    
    def redirect_to_appropriate_change(self, request, object_id):
        """Redirect to the appropriate change page based on record type"""
        try:
            obj = self.get_object(request, object_id)
            if obj:
                if obj.record_type == 'Order' and obj.order_item_id:
                    return redirect(reverse('admin:orders_orderitem_change', args=[obj.order_item_id]))
                elif obj.record_type == 'Reservation' and obj.reservation_id:
                    return redirect(reverse('admin:reservations_reservation_change', args=[obj.reservation_id]))
        except:
            pass
        # Fallback to default behavior
        return redirect(reverse('admin:orders_orderitemsmerged_changelist'))
    
    def get_record_type(self, obj):
        """Make the record_type field clickable to navigate to appropriate admin page"""
        if obj.record_type == 'Order' and obj.order_item_id:
            # Use order_item_id directly (no prefix in the current view)
            url = reverse('admin:orders_orderitem_change', args=[obj.order_item_id])
            return format_html('<a href="{}">{}</a>', url, obj.record_type)
        elif obj.record_type == 'Reservation' and obj.reservation_id:
            url = reverse('admin:reservations_reservation_change', args=[obj.reservation_id])
            return format_html('<a href="{}">{}</a>', url, obj.record_type)
        return obj.record_type
    
    get_record_type.short_description = 'Record Type'
    
    def get_party_size(self, obj):
        """
        Return party size with same logic as OrderItemAdmin:
        - For DINE_IN (orders and reservations): use party_size
        - For TAKEOUT orders: return 1
        - Reservations now have default order_type 'DINE_IN' from the view
        """
        try:
            if obj.order_type == 'DINE_IN':
                return obj.party_size if obj.party_size is not None else 'N/A'
            elif obj.order_type == 'TAKEOUT':
                return 1
            else:
                # Fallback for any edge cases
                return obj.party_size if obj.party_size is not None else 'N/A'
        except Exception:
            pass
        return 'N/A'
    
    get_party_size.short_description = 'Party Size'

    def get_user(self, obj):
        try:
            # Try to use the new direct user_email field from the view first
            if obj.user_email:
                return obj.user_email
            # Fallback to old method if user_email not available
            elif obj.user_id:
                from django.conf import settings
                from django.apps import apps
                User = apps.get_model(settings.AUTH_USER_MODEL)
                user = User.objects.filter(id=obj.user_id).first()
                return user if user else 'N/A'
        except Exception:
            pass
        return 'N/A'
    
    get_user.short_description = 'User'

    def get_venue_name(self, obj):
        try:
            # Try to use the new direct venue_name field from the view first
            if obj.venue_name:
                return obj.venue_name
            # Fallback to old method if venue_name not available
            elif obj.venue_id:
                from omochi.venues.models import Venue
                venue = Venue.objects.filter(id=obj.venue_id).first()
                return venue.name if venue else 'N/A'
        except Exception:
            pass
        return 'N/A'
    
    get_venue_name.short_description = 'Venue Name'

    def get_order_link(self, obj):
        # Get order link for either order or reservation
        from omochi.ref_logs.models import RefLog
        try:
            action_id = obj.order_id if obj.order_id else obj.reservation_id
            action_type = 'ORDER' if obj.order_id else 'RESERVATION'
            
            if action_id:
                ref_log = RefLog.objects.filter(action_id=str(action_id), action_type=action_type).first()
                if ref_log:
                    return ref_log.ref_id
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_order_link.short_description = 'Order Link'
    
    def get_ref_source_user_id(self, obj):
        # Get referral source user ID
        from omochi.ref_logs.models import RefLog
        from django.conf import settings
        from django.apps import apps
        
        try:
            action_id = obj.order_id if obj.order_id else obj.reservation_id
            action_type = 'ORDER' if obj.order_id else 'RESERVATION'
            
            if action_id:
                ref_log = RefLog.objects.filter(action_id=str(action_id), action_type=action_type).first()
                if ref_log and ref_log.ref_id:
                    User = apps.get_model(settings.AUTH_USER_MODEL)
                    referrer = User.objects.filter(ref_code=ref_log.ref_id).first()
                    if referrer:
                        return str(referrer.id)
            return 'None'
        except Exception as e:
            return f'Error: {str(e)}'
    
    get_ref_source_user_id.short_description = 'Ref Source User ID'

    def get_menu_item_name(self, obj):
        # Get menu item name from menu_item_id
        try:
            if obj.menu_item_id:
                from omochi.menus.models import MenuItem
                menu_item = MenuItem.objects.filter(id=obj.menu_item_id).first()
                return menu_item.name if menu_item else 'N/A'
        except Exception:
            pass
        return 'N/A'
    
    get_menu_item_name.short_description = 'Menu Item'

    def get_menu_price(self, obj):
        # Get menu price with takeout logic
        try:
            if obj.menu_item_id:
                from omochi.menus.models import MenuItem
                menu_item = MenuItem.objects.filter(id=obj.menu_item_id).first()
                if menu_item:
                    # Return take_out_price if order type is TAKEOUT and take_out_price exists
                    if obj.order_type == 'TAKEOUT' and hasattr(menu_item, 'take_out_price') and menu_item.take_out_price is not None:
                        return menu_item.take_out_price
                    else:
                        return menu_item.price
        except Exception:
            pass
        return 'N/A'
            
    get_menu_price.short_description = 'Unit Price'

    def get_order_date_jst(self, obj):
        # JST timezone
        jst = pytz.timezone('Asia/Tokyo')
        
        # Use order_date which now handles both order_date and reservation created_at
        if obj.order_date:
            # Convert UTC to JST
            utc_time = timezone.localtime(obj.order_date, pytz.UTC)
            jst_time = utc_time.astimezone(jst)
            return jst_time.strftime('%Y-%m-%d %H:%M:%S')
        return 'N/A'
    
    get_order_date_jst.short_description = 'Date of Order'

    def get_time_slot(self, obj):
        if obj.start_time and obj.end_time:
            return f"{obj.start_time} - {obj.end_time}"
        return 'N/A'
    
    get_time_slot.short_description = 'Time Slot'


@admin.register(OrderQuestion)
class OrderQuestionAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'order_link',
        'question_preview',
        'answer_preview',
        'order_index',
        'created_at',
    )
    list_filter = (
        'created_at',
        'order__venue',
    )
    search_fields = (
        'order__order_code',
        'question',
        'answer',
    )
    readonly_fields = (
        'id',
        'created_at',
        'updated_at',
    )
    ordering = ('order', 'order_index')
    
    def order_link(self, obj):
        """Create a link to order detail page"""
        if obj.order:
            url = reverse('admin:orders_order_change', args=[obj.order.id])
            return format_html('<a href="{}">{}</a>', url, obj.order.order_code)
        return "-"
    order_link.short_description = 'Order'
    order_link.admin_order_field = 'order__order_code'
    
    def question_preview(self, obj):
        """Display a preview of the question"""
        return obj.question[:50] + "..." if len(obj.question) > 50 else obj.question
    question_preview.short_description = 'Question'
    
    def answer_preview(self, obj):
        """Display a preview of the answer"""
        return obj.answer[:50] + "..." if len(obj.answer) > 50 else obj.answer
    answer_preview.short_description = 'Answer'
    
    def get_queryset(self, request):
        """Optimize queryset with select_related for better performance"""
        return super().get_queryset(request).select_related('order')
