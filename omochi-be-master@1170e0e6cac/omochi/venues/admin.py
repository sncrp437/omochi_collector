from django.contrib import admin
from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse
from openpyxl.styles import Alignment
from io import BytesIO
import openpyxl
import pytz

from omochi.common.openai_service import OpenAIService, get_openai_service
from .models import StockedVenue, Venue, VenueManager, VenueQuestion
from .forms import VenueAdminForm, VenueQuestionAdminForm


class VenueManagerInline(admin.TabularInline):
    model = VenueManager
    extra = 1
    autocomplete_fields = ['user']


class StockedVenueInline(admin.TabularInline):
    model = StockedVenue
    extra = 0
    readonly_fields = ['date_added']
    autocomplete_fields = ['user']


class VenueQuestionInline(admin.TabularInline):
    model = VenueQuestion
    extra = 1
    fields = ['question', 'question_en', 'ordinal']
    ordering = ['ordinal', 'created_at']
    verbose_name = "Question"
    verbose_name_plural = "Order Questions"
    
    def get_queryset(self, request):
        # Show questions ordered by ordinal field
        return super().get_queryset(request).order_by('ordinal', 'created_at')


@admin.register(Venue)
class VenueAdmin(admin.ModelAdmin):
    form = VenueAdminForm
    list_display = (
        'name',
        'name_en',
        'address',
        'nearest_station',
        'phone_number',
        'email',
        'is_partner',
        'opening_time',
        'closing_time',
        'stripe_account_status',
        'charges_enabled',
        'created_at',
        'updated_at'
    )
    list_filter = (
        'is_partner', 
        'nearest_station',
        'enable_reservation', 
        'enable_eat_in', 
        'enable_take_out',
        'genre', 
        'stripe_account_status',
        'charges_enabled',
        'payout_enabled',
    )
    search_fields = (
        'name', 'name_en', 'address', 'address_en', 
        'phone_number', 'email', 'genre', 'genre_en', 'stripe_account_id',
        'nearest_station'
    )
    readonly_fields = (
        'stripe_account_status', 'onboarding_complete',
        'charges_enabled', 'payout_enabled', 'created_at', 'updated_at'
    )
    fieldsets = (
        (None, {
            'fields': (
                ('name', 'name_en'),
                ('description', 'description_en'), 
                ('announcement', 'announcement_en'),
                'is_partner', 
                ('genre', 'genre_en')
            )
        }),
        (
            'Contact Information',
            {'fields': (
                ('address', 'address_en'),
                'phone_number', 
                'email', 
                'website',
                ('nearest_station', 'nearest_station_en')
            )},
        ),
        (
            'Features',
            {
                'fields': (
                    'opening_time',
                    'closing_time',
                    'additional_info',
                    'enable_cash_payment',
                    'enable_online_payment',
                    'enable_reservation',
                    'enable_eat_in',
                    'enable_take_out',
                )
            },
        ),
        (
            'Enable Order Questions',
            {
                'fields': ('enable_order_questions',),
                'description': 'Enable custom questions for customer orders. When enabled, you can add questions in the "Order Questions" section below.'
            },
        ),
        ('Media', {'fields': ('logo',)}),
        (
            'Stripe Connect',
            {
                'fields': (
                    'stripe_account_id',
                    'stripe_account_status',
                    'onboarding_complete',
                    'charges_enabled',
                    'payout_enabled',
                    'custom_platform_fee_amount',
                )
            },
        ),
        (
            'Timestamps',
            {'fields': ('created_at', 'updated_at'), 'classes': ('collapse',)},
        ),
    )
    inlines = [VenueQuestionInline, VenueManagerInline, StockedVenueInline]
    
    # Use custom template for change form
    change_form_template = 'admin/venues_venue_change_form.html'
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Override change view to handle translation requests"""
        # Clean up old session data to prevent memory leaks
        self._cleanup_old_session_data(request)
        
        if request.method == 'POST' and 'translate_action' in request.POST:
            # Check if this is an AJAX request
            if request.POST.get('translate_action') == 'ajax':
                return self._handle_ajax_translation_request(request, object_id)
            else:
                return self._handle_translation_request(request, object_id)
        
        return super().change_view(request, object_id, form_url, extra_context)

    def _cleanup_old_session_data(self, request):
        """Clean up old session data to prevent memory leaks"""
        keys_to_remove = []
        for key in request.session.keys():
            if key.startswith('translation_result_') or key == 'preserved_form_data':
                keys_to_remove.append(key)
        
        for key in keys_to_remove:
            del request.session[key]

    def _handle_ajax_translation_request(self, request, object_id):
        """Handle AJAX translation request without page reload"""
        from django.http import JsonResponse
        
        try:
            field_name = request.POST.get('translate_field')
            text = request.POST.get('translate_text', '').strip()
            
            # Validate input
            if not field_name:
                return JsonResponse({
                    'success': False,
                    'error': 'Field name is required'
                })
            
            # Limit text length to prevent abuse
            if len(text) > 5000:
                return JsonResponse({
                    'success': False,
                    'error': 'Text too long (max 5000 characters)'
                })
            
            # Handle potential URL decoding if needed
            try:
                import urllib.parse
                if '%' in text:
                    decoded_text = urllib.parse.unquote(text)
                    if decoded_text != text and len(decoded_text) > 0:
                        text = decoded_text
            except Exception:
                pass
            
            if not text:
                return JsonResponse({
                    'success': False,
                    'error': 'No text to translate'
                })
            
            # Initialize OpenAI service with timeout
            try:
                openai_service = OpenAIService()
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': 'Translation service unavailable'
                })
            
            # Call OpenAI translation
            result = openai_service.translate_text(
                text=text,
                target_language='English',
                source_language='Japanese',
                field_type=field_name,
                context="venue order question"
            )
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'translated_text': result['translated_text']
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Translation failed')
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Translation error: {str(e)}'
            })

    def _handle_translation_request(self, request, object_id):
        """Handle server-side translation without AJAX"""
        try:
            venue = self.get_object(request, object_id)
            if not venue:
                messages.error(request, "Venue not found")
                return self._redirect_back(request, object_id)
            
            field_name = request.POST.get('translate_field')
            text = request.POST.get('translate_text', '').strip()
            inline_field_name = request.POST.get('translate_inline_field')  # For inline forms
            
            # Preserve ALL form data (not just venue questions) to prevent data loss during page reload
            preserved_form_data = self._preserve_all_form_data(request)
            request.session['preserved_form_data'] = preserved_form_data
            
            # Handle potential URL decoding if needed
            try:
                import urllib.parse
                # Try to decode if it looks like URL encoded
                if '%' in text:
                    decoded_text = urllib.parse.unquote(text)
                    # Only use decoded if it's different and valid
                    if decoded_text != text and len(decoded_text) > 0:
                        text = decoded_text
            except Exception:
                # If decoding fails, use original text
                pass
            
            if not text:
                messages.error(request, "No text to translate")
                return self._redirect_back(request, object_id)
            
            # Initialize OpenAI service
            openai_service = OpenAIService()
            
            # Call OpenAI translation with field_type
            result = openai_service.translate_text(
                text=text,
                target_language='English',
                source_language='Japanese',
                field_type=field_name,
                context="venue order question" if field_name == 'question' else "venue information"
            )
            
            if result['success']:
                if inline_field_name:
                    # This is an inline translation - store result in session with specific field name
                    request.session[f'translation_result_{inline_field_name}'] = result['translated_text']
                    messages.success(
                        request, 
                        f"Successfully translated question to English: {result['translated_text'][:50]}..."
                    )
                else:
                    # This is a regular Venue field translation
                    setattr(venue, field_name, text)  # Save the Japanese text user entered
                    
                    en_field = f"{field_name}_en"
                    if hasattr(venue, en_field):
                        setattr(venue, en_field, result['translated_text'])  # Save translated English
                        venue.save()
                        
                        messages.success(
                            request, 
                            f"Successfully translated '{field_name}' to English: {result['translated_text'][:50]}..."
                        )
                    else:
                        messages.error(request, f"English field '{en_field}' not found")
            else:
                messages.error(request, f"Translation failed: {result.get('error', 'Unknown error')}")
                
        except Exception as e:
            messages.error(request, f"Translation error: {str(e)}")
        
        return self._redirect_back(request, object_id)
    
    def _preserve_all_form_data(self, request):
        """
        Preserve ALL form data during translation to prevent any data loss.
        
        This ensures that the entire form state is maintained when the page reloads,
        including all newly added inline forms and their data.
        """
        preserved_data = {}
        
        # Limit the size of preserved data to prevent memory issues
        max_field_size = 10000  # 10KB per field
        max_total_size = 100000  # 100KB total
        total_size = 0
        
        for field_name, field_value in request.POST.items():
            # Skip only translation-specific and security fields
            if field_name.startswith('translate_') or field_name == 'csrfmiddlewaretoken':
                continue
            
            # Check field size limits
            field_value_str = str(field_value)
            if len(field_value_str) > max_field_size:
                field_value_str = field_value_str[:max_field_size]
            
            total_size += len(field_value_str)
            if total_size > max_total_size:
                break
            
            # Preserve everything else to maintain complete form state
            preserved_data[field_name] = field_value_str
        
        return preserved_data
    
    def _redirect_back(self, request, object_id):
        """Redirect back to the change form"""
        from django.shortcuts import redirect
        return redirect('admin:venues_venue_change', object_id)


@admin.register(VenueManager)
class VenueManagerAdmin(admin.ModelAdmin):
    list_display = ('user', 'venue', 'role', 'created_at')
    list_filter = ('role', 'created_at')
    search_fields = ('user__email', 'venue__name')
    autocomplete_fields = ['user', 'venue']


@admin.register(StockedVenue)
class StockedVenueAdmin(admin.ModelAdmin):
    # Constants for Excel export
    EXCEL_COLUMN_PADDING = 2
    EXCEL_MAX_COLUMN_WIDTH = 75
    JST = pytz.timezone('Asia/Tokyo')
    
    list_display = (
        'user_email', 
        'user_full_name',
        'user_phone',
        'user_address',
        'venue_name', 
        'is_favorite', 
        'get_date_added_jst'
    )
    list_filter = ('is_favorite', 'date_added')
    search_fields = (
        'user__email', 
        'user__first_name', 
        'user__last_name',
        'user__phone_number',
        'user__addresses__prefecture',
        'user__addresses__city',
        'user__addresses__detail',
        'venue__name'
    )
    autocomplete_fields = ['user', 'venue']
    readonly_fields = ['date_added']
    ordering = ['-date_added']
    actions = ['export_selected_to_excel', 'export_all_to_excel']
    
    # Disable clickable links to detail view
    list_display_links = None
    
    def get_queryset(self, request):
        """Optimize queryset to prevent N+1 queries"""
        return super().get_queryset(request).select_related(
            'user', 
            'venue'
        ).prefetch_related(
            'user__addresses'
        )
    
    def changelist_view(self, request, extra_context=None):
        """Override to handle export_all action without requiring selection"""
        if 'action' in request.POST and request.POST['action'] == 'export_all_to_excel':
            # Skip selection check for export_all action
            if not request.POST.getlist('_selected_action'):
                # Get the changelist to access the filtered queryset
                cl = self.get_changelist_instance(request)
                # Get filtered queryset (respects search and filters)
                filtered_queryset = cl.get_queryset(request)
                # Call the action with filtered queryset
                return self.export_all_to_excel(request, filtered_queryset)
        return super().changelist_view(request, extra_context)
    
    def has_add_permission(self, request):
        """Make the admin read-only by disabling add permission"""
        return False
    
    def has_change_permission(self, request, obj=None):
        """Make the admin read-only by disabling change permission"""
        return False
    
    def has_delete_permission(self, request, obj=None):
        """Make the admin read-only by disabling delete permission"""
        return False
    
    def user_email(self, obj):
        """Display user email"""
        return obj.user.email if obj.user else '-'
    user_email.short_description = 'User Email'
    user_email.admin_order_field = 'user__email'
    
    def user_full_name(self, obj):
        """Display user full name"""
        if obj.user:
            full_name = f"{obj.user.first_name} {obj.user.last_name}".strip()
            return full_name if full_name else obj.user.email
        return '-'
    user_full_name.short_description = 'User Name'
    user_full_name.admin_order_field = 'user__first_name'
    
    def user_phone(self, obj):
        """Display user phone number"""
        return getattr(obj.user, 'phone_number', '-') if obj.user else '-'
    user_phone.short_description = 'Phone Number'
    
    def _get_user_default_address(self, user, full=False):
        """Get user's default address or first available address"""
        if not user or not user.addresses.exists():
            return None
        
        default_address = user.addresses.filter(is_default=True).first()
        if not default_address:
            default_address = user.addresses.first()
        
        if not default_address:
            return None
            
        if full:
            return f"{default_address.prefecture} {default_address.city} {default_address.detail}"
        return f"{default_address.prefecture} {default_address.city}"
    
    def user_address(self, obj):
        """Display user address"""
        address = self._get_user_default_address(obj.user, full=False)
        return address if address else '-'
    user_address.short_description = 'User Address'
    
    def venue_name(self, obj):
        """Display venue name"""
        return obj.venue.name if obj.venue else '-'
    venue_name.short_description = 'Venue Name'
    venue_name.admin_order_field = 'venue__name'
    
    def venue_address(self, obj):
        """Display venue address"""
        return obj.venue.address if obj.venue else '-'
    venue_address.short_description = 'Venue Address'
    venue_address.admin_order_field = 'venue__address'
    
    def get_date_added_jst(self, obj):
        if obj.date_added:
            # Direct conversion from UTC to JST
            jst_time = obj.date_added.astimezone(self.JST)
            return jst_time.strftime('%Y-%m-%d %H:%M:%S')
        return '-'
    get_date_added_jst.short_description = 'Date Added (JST)'
    get_date_added_jst.admin_order_field = 'date_added'
    
    def _export_to_excel_helper(self, queryset):
        """Helper method to generate Excel file from queryset"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stocked Venues"
        
        # Write headers
        headers = [
            'User Email',
            'User Name', 
            'User Phone',
            'User Address',
            'Venue Name',
            'Date Added'
        ]
        ws.append(headers)
        
        # Format header row
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for obj in queryset:
            user = obj.user
            venue = obj.venue
            
            # Get user's default address using helper method
            user_address = self._get_user_default_address(user, full=True)
            if not user_address:
                user_address = 'N/A'
            
            # Get date_added in JST timezone using the same method as list display
            date_added_jst = self.get_date_added_jst(obj)
            if date_added_jst == '-':
                date_added_jst = 'N/A'
            
            # Prepare row data (avoid unnecessary str() conversions)
            row = [
                user.email if user else 'N/A',
                f"{user.first_name} {user.last_name}".strip() if user else 'N/A',
                user.phone_number if user and user.phone_number else 'N/A',
                user_address,
                venue.name if venue else 'N/A',
                date_added_jst
            ]
            
            ws.append(row)
        
        # Auto-adjust column widths to fit content
        for col_idx, column_cells in enumerate(ws.iter_cols(min_col=1, max_col=len(headers)), start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    # Handle cases where cell.value cannot be converted to string
                    pass
            # Set column width with some padding (max 75 to prevent extremely wide columns)
            adjusted_width = min(max_length + self.EXCEL_COLUMN_PADDING, self.EXCEL_MAX_COLUMN_WIDTH)
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Use context manager to ensure buffer is properly closed
        with BytesIO() as buffer:
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                content=buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return response
    
    def export_selected_to_excel(self, request, queryset):
        """Export only selected stocked venues to Excel"""
        response = self._export_to_excel_helper(queryset)
        response['Content-Disposition'] = 'attachment; filename=stocked_venues_selected.xlsx'
        return response
    
    export_selected_to_excel.short_description = "Export Selected Stocked Venues to Excel"
    
    def export_all_to_excel(self, request, queryset):
        """Export ALL stocked venues based on current filters/search (ignoring pagination)"""
        # Use the filtered queryset passed from changelist_view
        # This respects all current filters and search
        if queryset is None:
            queryset = self.get_queryset(request)
        
        response = self._export_to_excel_helper(queryset)
        response['Content-Disposition'] = 'attachment; filename=stocked_venues_filtered.xlsx'
        return response
    
    export_all_to_excel.short_description = "Export ALL Filtered Results to Excel"


@admin.register(VenueQuestion)
class VenueQuestionAdmin(admin.ModelAdmin):
    form = VenueQuestionAdminForm
    list_display = ('question', 'question_en', 'venue', 'ordinal', 'created_at')
    list_filter = ('created_at', 'venue')
    search_fields = ('question', 'question_en', 'venue__name')
    autocomplete_fields = ['venue']
    ordering = ['venue', 'ordinal', 'created_at']
    
    # Use custom template for change form  
    change_form_template = 'admin/venues_venuequestion_change_form.html'
    
    fieldsets = (
        (None, {
            'fields': ('venue', ('question', 'question_en'),)
        }),
        ('Settings', {
            'fields': ('ordinal',)
        }),
    )

    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Override change view to handle translation requests"""
        if request.method == 'POST' and 'translate_action' in request.POST:
            return self._handle_translation_request(request, object_id)
        
        return super().change_view(request, object_id, form_url, extra_context)

    def _handle_translation_request(self, request, object_id):
        """Handle server-side translation without AJAX"""
        try:
            venue_question = self.get_object(request, object_id)
            if not venue_question:
                messages.error(request, "Venue question not found")
                return self._redirect_back(request, object_id)

            field_name = request.POST.get('translate_field')
            text = request.POST.get('translate_text', '').strip()

            # Handle potential URL decoding if needed
            try:
                import urllib.parse
                # Try to decode if it looks like URL encoded
                if '%' in text:
                    decoded_text = urllib.parse.unquote(text)
                    # Only use decoded if it's different and valid
                    if decoded_text != text and len(decoded_text) > 0:
                        text = decoded_text
            except Exception:
                # If decoding fails, use original text
                pass

            if not text:
                messages.error(request, "No text to translate")
                return self._redirect_back(request, object_id)

            # Initialize OpenAI service
            openai_service = get_openai_service()

            # Call OpenAI translation with field_type
            result = openai_service.translate_text(
                text=text,
                target_language='English',
                source_language='Japanese',
                field_type=field_name,
                context="venue order question"
            )

            if result['success']:
                # Update BOTH the source field and corresponding English field
                # This preserves the user's input data
                setattr(venue_question, field_name, text)  # Save the Japanese text user entered
                
                en_field = f"{field_name}_en"
                if hasattr(venue_question, en_field):
                    setattr(venue_question, en_field, result['translated_text'])  # Save translated English
                    venue_question.save()
                    
                    messages.success(
                        request, 
                        f"Successfully translated '{field_name}' to English: {result['translated_text'][:50]}..."
                    )
                else:
                    messages.error(request, f"English field '{en_field}' not found")
            else:
                messages.error(request, f"Translation failed: {result.get('error', 'Unknown error')}")

        except Exception as e:
            messages.error(request, f"Translation error: {str(e)}")

        return self._redirect_back(request, object_id)

    def _redirect_back(self, request, object_id):
        """Redirect back to the change form"""
        from django.urls import reverse
        return redirect(reverse(
            f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_change', 
            args=[object_id]
        ))